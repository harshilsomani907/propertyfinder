# ==============================================================================
#  PROPERTY FINDER AUTOMATED INCREMENTAL DETAIL SCRAPER
#  Scrapes: Property details (Title, Price, Location, Beds, Baths, Area, Link,
#           Image Link, Description, Amenities, Property Survey No, City, City Area,
#           Purpose, Property Type, Furnishing, Price per Sqft, Verified, Listed On,
#           Parking Spaces, Agent Name, Agent Phone, Agent WhatsApp)
#  Output: Updates propertyfinder_detailed_properties.xlsx and dumps to json
# ==============================================================================

import os
import sys
import time
import json
import argparse
import requests
import re
import subprocess
import shutil
from io import BytesIO

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from PIL import Image as PILImage
    from openpyxl.drawing.image import Image as OpenpyxlImage
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False


def get_chrome_version():
    """Detects the installed Chrome/Chromium version on Windows, Linux, or macOS."""
    if os.name == 'nt':
        # Windows Registry Checks
        try:
            cmd = r'reg query "HKEY_CURRENT_USER\Software\Google\Chrome\BLBeacon" /v version'
            output = subprocess.check_output(cmd, shell=True).decode()
            version = re.search(r'version\s+REG_SZ\s+([\d\.]+)', output)
            if version:
                return version.group(1)
        except Exception:
            pass
        
        try:
            cmd = r'reg query "HKEY_LOCAL_MACHINE\Software\Wow6432Node\Microsoft\Windows\CurrentVersion\Uninstall\Google Chrome" /v version'
            output = subprocess.check_output(cmd, shell=True).decode()
            version = re.search(r'version\s+REG_SZ\s+([\d\.]+)', output)
            if version:
                return version.group(1)
        except Exception:
            pass

        # Check common executable paths version info on Windows
        paths = [
            os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe")
        ]
        for path in paths:
            if os.path.exists(path):
                try:
                    cmd = f'powershell -command "(Get-Item \'{path}\').VersionInfo.ProductVersion"'
                    output = subprocess.check_output(cmd, shell=True).decode().strip()
                    if output:
                        return output
                except Exception:
                    pass
    else:
        # Linux / macOS Checks
        for cmd in ['google-chrome', 'google-chrome-stable', 'chromium', 'chromium-browser']:
            try:
                output = subprocess.check_output([cmd, '--version']).decode()
                version = re.search(r'([\d\.]+)', output)
                if version:
                    return version.group(1)
            except Exception:
                continue
    return None


def find_chrome_executable():
    """Locates the path of the Chrome or Chromium binary."""
    if os.name == 'nt':
        paths = [
            os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe")
        ]
        for path in paths:
            if os.path.exists(path):
                return path
    else:
        # Check standard binary command names on Linux / macOS
        for cmd in ['google-chrome', 'google-chrome-stable', 'chromium', 'chromium-browser']:
            path = shutil.which(cmd)
            if path:
                return path
    return None


def find_chromedriver_executable():
    """Locates the system-installed chromedriver binary (critical for Nix environments)."""
    if os.name == 'nt':
        return None  # Let undetected-chromedriver handle it on Windows
    
    # Check PATH first
    path = shutil.which('chromedriver')
    if path:
        return path
    
    # Check common Nix / system locations
    common_paths = [
        '/usr/bin/chromedriver',
        '/usr/local/bin/chromedriver',
    ]
    for p in common_paths:
        if os.path.exists(p) and os.access(p, os.X_OK):
            return p
    
    # Search in /nix/store for chromedriver
    nix_store = '/nix/store'
    if os.path.isdir(nix_store):
        try:
            for entry in os.listdir(nix_store):
                candidate = os.path.join(nix_store, entry, 'bin', 'chromedriver')
                if os.path.exists(candidate) and os.access(candidate, os.X_OK):
                    return candidate
        except Exception:
            pass
    
    return None


def _make_chrome_options(is_headless):
    """Creates a fresh ChromeOptions instance (cannot be reused across uc.Chrome calls)."""
    options = uc.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    if is_headless:
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-software-rasterizer")
        options.add_argument("--single-process")
        options.add_argument("--disable-extensions")
    return options


def open_browser():
    is_headless = os.name != 'nt'
    if is_headless:
        print("🌐 Launching Chrome in headless mode using undetected-chromedriver...")
    else:
        print("🌐 Launching headed Chrome using undetected-chromedriver...")

    chrome_path = find_chrome_executable()
    chromedriver_path = find_chromedriver_executable()
    ver = get_chrome_version()
    version_main = None
    
    if ver:
        try:
            version_main = int(ver.split('.')[0])
            print(f"ℹ️ Detected Chrome version: {ver}. Forcing version_main={version_main}")
        except Exception as e:
            print(f"⚠️ Error parsing Chrome version '{ver}': {e}")

    if chrome_path:
        print(f"ℹ️ Using Chrome executable: {chrome_path}")
    else:
        print("⚠️ No Chrome executable found by find_chrome_executable()")
        
    if chromedriver_path:
        print(f"ℹ️ Using system chromedriver: {chromedriver_path}")
    else:
        print("⚠️ No system chromedriver found by find_chromedriver_executable()")
        print("⚠️ Nixpacks configuration might be missing the 'chromedriver' package, or it's not in PATH.")

    # CRITICAL FIX FOR NIXPACKS:
    # undetected-chromedriver will try to download a generic Ubuntu binary if it thinks
    # the system one is outdated, which causes exit code 127 on NixOS/Railway.
    # We monkeypatch the patcher to prevent downloading if we already have a system binary.
    if chromedriver_path and is_headless:
        print("🛡️ Monkeypatching undetected_chromedriver to prevent generic binary download...")
        import undetected_chromedriver.patcher
        undetected_chromedriver.patcher.Patcher.fetch_package = lambda *args, **kwargs: print("   -> Blocked fetch_package()")
        undetected_chromedriver.patcher.Patcher.unzip_package = lambda *args, **kwargs: print("   -> Blocked unzip_package()")

    # Build kwargs with all detected paths
    options = _make_chrome_options(is_headless)
    kwargs = {"options": options, "headless": is_headless}
    if chrome_path:
        kwargs["browser_executable_path"] = chrome_path
    if chromedriver_path:
        kwargs["driver_executable_path"] = chromedriver_path
    if version_main:
        kwargs["version_main"] = version_main
    
    try:
        print(f"🚀 Attempt 1: Launching uc.Chrome with kwargs: {kwargs}")
        driver = uc.Chrome(**kwargs)
    except Exception as err:
        print(f"⚠️ Launch failed: {err}")
        print("🔄 Retrying with fresh options and no custom driver path...")
        options2 = _make_chrome_options(is_headless)
        kwargs2 = {"options": options2, "headless": is_headless}
        if chrome_path:
            kwargs2["browser_executable_path"] = chrome_path
        if version_main:
            kwargs2["version_main"] = version_main
        # Don't pass driver_executable_path - let it download its own
        print(f"🚀 Attempt 2: Launching uc.Chrome with kwargs: {kwargs2}")
        driver = uc.Chrome(**kwargs2)
        
    try:
        if not is_headless:
            driver.maximize_window()
    except Exception:
        pass
    return driver


def load_existing_links(file_path):
    """Loads existing listing URLs from Excel sheet to avoid duplicate scraping."""
    if not os.path.exists(file_path):
        print(f"ℹ️ Output Excel file does not exist yet. Will create a new one.")
        return set()
        
    print(f"📖 Reading existing links from {file_path} (read-only mode)...")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_name = 'Property Finder Listings'
        if sheet_name not in wb.sheetnames:
            print(f"⚠️ Sheet '{sheet_name}' not found. Loading active sheet.")
            ws = wb.active
        else:
            ws = wb[sheet_name]
            
        links = set()
        # Col 9 (I) is Link column
        for row in ws.iter_rows(min_row=2, min_col=9, max_col=9, values_only=True):
            val = row[0]
            if val and str(val).startswith("http"):
                links.add(str(val).strip())
                
        wb.close()
        print(f"✅ Loaded {len(links)} existing links from spreadsheet.")
        return links
    except Exception as e:
        print(f"⚠️ Failed to read existing Excel file: {e}. Starting fresh.")
        return set()


def scrape_details(driver, url):
    """Navigates to detail page, extracts __NEXT_DATA__ JSON tag, and parses all 25 fields."""
    print(f"🔗 Navigating to details: {url}")
    try:
        driver.get(url)
        # Give page some time to render and bypass challenge
        time.sleep(4)
        
        # Wait for NEXT_DATA tag
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "__NEXT_DATA__"))
        )
        
        script_el = driver.find_element(By.ID, "__NEXT_DATA__")
        json_text = script_el.get_attribute("innerHTML")
        data = json.loads(json_text)
        
        # Access nested details
        page_props = data.get("props", {}).get("pageProps", {})
        prop_res = page_props.get("propertyResult", {})
        p = prop_res.get("property", {})
        if not p:
            print("  ❌ Property details not found in NEXT_DATA JSON.")
            return None
            
        # Parse fields
        title = p.get("title", "N/A")
        
        # Price
        price_obj = p.get("price", {})
        price_val = price_obj.get("value", 0)
        price_period = price_obj.get("period", "")
        price_freq = f"/{price_period}" if price_period else ""
        price_str = f"AED {price_val:,}{price_freq}" if price_val else "N/A"
        
        # Location
        loc_obj = p.get("location", {})
        location_str = loc_obj.get("full_name", "N/A")
        
        # Specs
        beds = p.get("bedrooms", "N/A")
        baths = p.get("bathrooms", "N/A")
        
        size_obj = p.get("size", {})
        size_val = size_obj.get("value", 0)
        size_unit = size_obj.get("unit", "sqft")
        size_str = f"{size_val} {size_unit}" if size_val else "N/A"
        
        # Description & Amenities
        description = p.get("description", "N/A")
        amenities_list = p.get("amenities", [])
        amenities_str = ", ".join([a.get("name") for a in amenities_list if a.get("name")]) if amenities_list else "N/A"
        
        # Survey No (rera number)
        survey_no = p.get("rera", {}).get("number")
        if not survey_no:
            survey_no = p.get("reference", "N/A")
            
        # Location tree
        loc_tree = p.get("location_tree", [])
        city = "Dubai"
        city_area = "N/A"
        for item in loc_tree:
            t = item.get("type", "").upper()
            if t == "CITY":
                city = item.get("name", city)
            elif t == "COMMUNITY":
                city_area = item.get("name", city_area)
                
        # Purpose
        is_rent = page_props.get("isRent", True)
        purpose_str = "Rent" if is_rent else "Sale"
        
        # Property Type & Furnishing
        property_type = p.get("property_type", "N/A")
        furnished = p.get("furnished", "")
        furnishing_str = "Furnished" if furnished == "YES" or furnished is True else "Unfurnished" if furnished == "NO" or furnished is False else str(furnished) if furnished else "N/A"
        
        # Price per Sqft
        price_per_sqft = "N/A"
        if price_val and size_val:
            price_per_sqft = f"{price_val / size_val:.2f}"
            
        # Verified & Listed Date
        verified = "Yes" if p.get("is_verified") else "No"
        listed_on_raw = p.get("listed_date", "")
        listed_on = listed_on_raw[:10] if listed_on_raw else "N/A"
        
        # Parking
        parking_spaces = p.get("parking_spaces", "N/A")
        
        # Agent Details
        agent = p.get("agent", {})
        agent_name = agent.get("name", "N/A")
        
        agent_phone = "N/A"
        agent_whatsapp = "N/A"
        contact_options = p.get("contact_options", [])
        for opt in contact_options:
            opt_type = opt.get("type", "").lower()
            if opt_type == "phone":
                agent_phone = opt.get("value", "N/A")
            elif opt_type == "whatsapp":
                agent_whatsapp = opt.get("value", "N/A")
                
        # Image
        img_url = "N/A"
        images_list = p.get("images", {}).get("property", [])
        if images_list:
            img_url = images_list[0].get("full") or images_list[0].get("medium") or "N/A"
            
        # Validation url
        dubailand_url = p.get("rera", {}).get("permit_validation_url", "N/A")
        
        return {
            "Title": title,
            "Price": price_str,
            "Location": location_str,
            "Beds": beds,
            "Baths": baths,
            "Area": size_str,
            "Link": url,
            "Image_URL": img_url,
            "Description": description,
            "Amenities": amenities_str,
            "Property Survey No": survey_no,
            "City": city,
            "City Area": city_area,
            "Purpose": purpose_str,
            "Property Type": property_type,
            "Furnishing": furnishing_str,
            "Price per Sqft": price_per_sqft,
            "Verified": verified,
            "Listed On": listed_on,
            "Parking Spaces": parking_spaces,
            "Agent Name": agent_name,
            "Agent Phone": agent_phone,
            "Agent WhatsApp": agent_whatsapp,
            "Dubailand_Validation_URL": dubailand_url
        }
    except Exception as e:
        print(f"  ❌ Error scraping detail page: {e}")
        return None


def fetch_listing_cards(driver):
    """Waits for listing cards and returns their links."""
    card_selectors = [
        "article[data-testid='property-card']",
        "div[class*='property-card']",
        "article"
    ]
    
    cards = []
    for selector in card_selectors:
        elements = driver.find_elements(By.CSS_SELECTOR, selector)
        if elements and len(elements) > 2:
            cards = elements
            break
            
    links = []
    for card in cards:
        try:
            link_el = card.find_elements(By.CSS_SELECTOR, 'a[class*="property-card__link"], a[class*="card-link"]')
            if not link_el:
                link_el = card.find_elements(By.CSS_SELECTOR, 'a[href*="/en/plp/"], a[href*="/en/buy/"], a[href*="/en/rent/"]')
            if not link_el:
                link_el = card.find_elements(By.CSS_SELECTOR, 'a')
                
            if link_el:
                href = link_el[0].get_attribute("href")
                if href and href.startswith("http") and not "javascript" in href:
                    links.append(href)
        except Exception:
            continue
            
    # De-duplicate links while preserving order
    unique_links = []
    for l in links:
        if l not in unique_links:
            unique_links.append(l)
    return unique_links


def save_to_excel(new_listings, output_file):
    """Loads existing Excel workbook, appends new rows with image downloading, and saves."""
    print(f"\n💾 Updating Excel spreadsheet at {output_file}...")
    
    # Check if file exists, if not create new workbook
    if os.path.exists(output_file):
        try:
            wb = openpyxl.load_workbook(output_file)
        except Exception as e:
            print(f"⚠️ Excel file corrupted or unreadable ({e}). Recreating a new workbook.")
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()
        
    sheet_name = 'Property Finder Listings'
    if sheet_name not in wb.sheetnames:
        ws = wb.active
        ws.title = sheet_name
        
        # Create headers
        headers = [
            "#", "Photo", "Title", "Price", "Location", "Beds", "Baths", "Area", "Link", "Image Link",
            "Description", "Amenities", "Property Survey No", "City", "City Area", "Purpose",
            "Property Type", "Furnishing", "Price per Sqft", "Verified", "Listed On", "Parking Spaces",
            "Agent Name", "Agent Phone", "Agent WhatsApp"
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
    else:
        ws = wb[sheet_name]
        
    # Get current index offset
    current_rows = ws.max_row
    
    # Styling variables
    light_blue = PatternFill("solid", fgColor="F2F6F9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    link_font = Font(name="Segoe UI", color="0563C1", underline="single")
    default_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    for idx, prop in enumerate(new_listings, 1):
        row_num = current_rows + idx
        row_fill = light_blue if row_num % 2 == 0 else white_fill
        
        row_data = [
            row_num - 1, # Index #
            "",          # Photo placeholder
            prop.get("Title", "N/A"),
            prop.get("Price", "N/A"),
            prop.get("Location", "N/A"),
            prop.get("Beds", "N/A"),
            prop.get("Baths", "N/A"),
            prop.get("Area", "N/A"),
            prop.get("Link", "N/A"),
            prop.get("Image_URL", "N/A"),
            prop.get("Description", "N/A"),
            prop.get("Amenities", "N/A"),
            prop.get("Property Survey No", "N/A"),
            prop.get("City", "Dubai"),
            prop.get("City Area", "N/A"),
            prop.get("Purpose", "Rent"),
            prop.get("Property Type", "N/A"),
            prop.get("Furnishing", "N/A"),
            prop.get("Price per Sqft", "N/A"),
            prop.get("Verified", "N/A"),
            prop.get("Listed On", "N/A"),
            prop.get("Parking Spaces", "N/A"),
            prop.get("Agent Name", "N/A"),
            prop.get("Agent Phone", "N/A"),
            prop.get("Agent WhatsApp", "N/A")
        ]
        
        ws.append(row_data)
        
        # Formatting and styling
        ws.row_dimensions[row_num].height = 75
        for col_num, cell in enumerate(ws[row_num], 1):
            cell.fill = row_fill
            cell.font = default_font
            cell.border = thin_border
            
            # Alignments
            if col_num in (1, 6, 7, 8, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            # Links colors
            if col_num in (9, 10) and cell.value not in ("N/A", ""):
                cell.font = link_font
                
        # Embed photo
        img_url = prop.get("Image_URL", "N/A")
        if img_url != "N/A" and img_url.startswith("http") and HAS_PILLOW:
            try:
                print(f"📸 Downloading thumbnail for Row {row_num}...")
                resp = requests.get(img_url, timeout=5)
                if resp.status_code == 200:
                    pil_img = PILImage.open(BytesIO(resp.content))
                    if pil_img.mode in ("RGBA", "P"):
                        pil_img = pil_img.convert("RGB")
                    # Thumbnail sizing to fit row cleanly
                    pil_img.thumbnail((120, 90))
                    
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='JPEG')
                    img_byte_arr.seek(0)
                    
                    xlsx_img = OpenpyxlImage(img_byte_arr)
                    ws.add_image(xlsx_img, f"B{row_num}")
            except Exception as e:
                print(f"  ⚠️ Photo load failed: {e}")
                
    # Format widths of columns if they're not set
    column_widths = {
        "A": 6, "B": 18, "C": 35, "D": 18, "E": 30, "F": 8, "G": 8, "H": 12, "I": 25, "J": 25,
        "K": 40, "L": 30, "M": 18, "N": 12, "O": 15, "P": 10, "Q": 15, "R": 12, "S": 15, "T": 10,
        "U": 12, "V": 10, "W": 18, "X": 16, "Y": 16
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Setup summary sheet
    summary_sheet = "Summary"
    if summary_sheet in wb.sheetnames:
        ws_sum = wb[summary_sheet]
    else:
        ws_sum = wb.create_sheet(title=summary_sheet)
        
    ws_sum["A1"] = "PropertyFinder Autopilot Scraping Summary"
    ws_sum["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="1F4E79")
    
    ws_sum["A3"] = "Last Daily Run:"
    ws_sum["B3"] = time.strftime("%Y-%m-%d %H:%M:%S")
    ws_sum["A4"] = "Properties Scraped in Run:"
    ws_sum["B4"] = len(new_listings)
    ws_sum["A5"] = "Total Properties in Sheet:"
    ws_sum["B5"] = ws.max_row - 1
    
    ws_sum.column_dimensions["A"].width = 25
    ws_summary_width = 30
    ws_sum.column_dimensions["B"].width = ws_summary_width
    
    # Save safely to temporary file first
    temp_output = output_file + ".tmp"
    wb.save(temp_output)
    wb.close()
    
    if os.path.exists(output_file):
        os.remove(output_file)
    os.rename(temp_output, output_file)
    print(f"🎉 Spreadsheet saved successfully: {output_file}")


def main():
    parser = argparse.ArgumentParser(description="Scrape details of PropertyFinder Dubai Rent listings.")
    parser.add_argument("--pages", type=int, default=3, help="Pages to scan")
    parser.add_argument("--output", default="propertyfinder_detailed_properties.xlsx", help="Spreadsheet output path")
    parser.add_argument("--new-json", default="new_listings.json", help="Path to write new listings JSON data")
    
    args = parser.parse_args()
    
    print("=" * 65)
    print("     PROPERTYFINDER INCREMENTAL AUTO-RUN SCRAPER")
    print("=" * 65)
    print(f" Pages depth: {args.pages}")
    print(f" Output File: {args.output}")
    print(f" JSON Dump  : {args.new_json}")
    print("=" * 65)
    
    # Resolve absolute paths
    abs_output = os.path.abspath(args.output)
    abs_json = os.path.abspath(args.new_json)
    
    # Load already scraped links to bypass details fetching
    scraped_links = load_existing_links(abs_output)
    
    driver = open_browser()
    new_details = []
    
    try:
        for page in range(1, args.pages + 1):
            print(f"\n📄 Scanning search page {page} of {args.pages}...")
            
            if page == 1:
                url = "https://www.propertyfinder.ae/en/rent/dubai/properties-for-rent.html?ob=nd"
            else:
                url = f"https://www.propertyfinder.ae/en/rent/dubai/properties-for-rent.html?page={page}&ob=nd"
                
            driver.get(url)
            time.sleep(5)
            
            links = fetch_listing_cards(driver)
            print(f"  Found {len(links)} property links in search page.")
            
            for index, link in enumerate(links):
                # Unique ID checking
                if link in scraped_links:
                    print(f"  ⏩ [{index+1}/{len(links)}] Skipping already scraped: {link}")
                    continue
                    
                print(f"  🚀 [{index+1}/{len(links)}] Scraping details for new listing...")
                time.sleep(2) # Politeness delay
                
                details = scrape_details(driver, link)
                if details:
                    new_details.append(details)
                    # Add to scraped set dynamically to prevent duplicates
                    scraped_links.add(link)
                    print(f"  ✅ Extracted: {details['Title']} ({details['Price']})")
                
                # Add larger rest between detail pages
                time.sleep(3)
                
    except KeyboardInterrupt:
        print("\n⚠️ Scraper run aborted by keyboard interrupt.")
    finally:
        driver.quit()
        print("\n🔒 Browser closed.")
        
    print(f"\n📊 Run Completed. Scraped {len(new_details)} new listings.")
    
    if new_details:
        # Save JSON output for the server to process
        with open(abs_json, "w", encoding="utf-8") as f:
            json.dump(new_details, f, indent=2, ensure_ascii=False)
        print(f"💾 Scraped data dumped to {abs_json}")
        
        # Update the spreadsheet
        save_to_excel(new_details, abs_output)
    else:
        print("ℹ️ No new listings found to update.")
        # Make sure empty list json is written so server doesn't get blocked
        with open(abs_json, "w", encoding="utf-8") as f:
            json.dump([], f)


if __name__ == "__main__":
    main()
